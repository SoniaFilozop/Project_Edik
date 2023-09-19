import os
import sys

import win32com.client
from PIL import ImageGrab
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QFont, QColor
from PyQt5.QtWidgets import QApplication, QLabel, QMainWindow, QFileDialog, QWidget, QPushButton, QLineEdit, \
    QDesktopWidget, QGraphicsDropShadowEffect
import xlsxwriter
from PyQt5.uic.properties import QtCore
from docx.shared import Cm
from docxtpl import DocxTemplate, InlineImage
from docx.shared import Cm, Inches, Mm, Emu
import random
import datetime
from openpyxl import load_workbook
import openpyxl as xl

SCREEN_SIZE = [400, 400]


class Example(QWidget):

    def __init__(self):
        super().__init__()
        self.initUI()

    def center(self):

        qr = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())

    def initUI(self):
        self.setGeometry(100, 400, 700, 550)
        self.move(550, 200)
        self.center()
        self.setWindowTitle('Фильтр таблиц')
        self.setStyleSheet("background-color:#3b2530")
        self.setStyleSheet("background-color:black")
        self.pushButton = QPushButton('Отфильтровать таблицу', self)
        self.pushButton.setFont(QFont('Calibri', 20))
        self.pushButton.setStyleSheet('background-color:white; border-radius: 20px')
        effect = QGraphicsDropShadowEffect(self)
        effect.setBlurRadius(12)
        effect.setOffset(4, 4)
        effect.setColor(Qt.gray)
        self.pushButton.setGraphicsEffect(effect)
        self.pushButton.resize(400, 90)
        self.pushButton.clicked.connect(self.run)
        self.pushButton.move(150, 400)
        self.lineEdit = QLineEdit('file.xlsx', self)
        self.lineEdit.setFont(QFont('Calibri', 13))
        self.lineEdit.setStyleSheet('color:white')
        self.lineEdit.resize(200, 40)
        self.lineEdit.move(450, 150)
        self.lineEdit_2 = QLineEdit('130025', self)
        self.lineEdit_2.setFont(QFont('Calibri', 13))
        self.lineEdit_2.setStyleSheet('color:white')
        self.lineEdit_2.resize(200, 40)
        self.lineEdit_2.move(450, 75)
        self.lineEdit_3 = QLineEdit(self)
        self.lineEdit_3.setFont(QFont('Calibri', 13))
        self.lineEdit_3.setStyleSheet('color:white')
        self.lineEdit_3.resize(200, 40)
        self.lineEdit_3.move(450, 225)
        self.lineEdit_4 = QLineEdit('file2.xlsx', self)
        self.lineEdit_4.setFont(QFont('Calibri', 13))
        self.lineEdit_4.setStyleSheet('color:white')
        self.lineEdit_4.resize(200, 40)
        self.lineEdit_4.move(450, 300)
        self.label = QLabel('Имя, под которым сохранить отфильтрованный файл', self)
        self.label.setFont(QFont('Calibri', 13))
        self.label.setStyleSheet('color:white')
        self.label.resize(self.label.sizeHint())
        self.label.move(20, 160)
        self.label_2 = QLabel('Код образовательной организации', self)
        self.label_2.setFont(QFont('Calibri', 13))
        self.label_2.setStyleSheet('color:white')
        self.label_2.resize(self.label.sizeHint())
        self.label_2.move(20, 85)
        self.label_3 = QLabel('Класс', self)
        self.label_3.setFont(QFont('Calibri', 13))
        self.label_3.setStyleSheet('color:white')
        self.label_3.resize(200, 20)
        self.label_3.move(20, 235)
        self.label_4 = QLabel('Имя, под которым сохранить файл с выбранным классом', self)
        self.label_4.setFont(QFont('Calibri', 13))
        self.label_4.setStyleSheet('color:white')
        self.label_4.resize(420, 20)
        self.label_4.move(20, 310)

    def make_list(self, s, d, v, sm, n, all):
        # Workbook () принимает один необязательный аргумент
        # это имя файла, которое мы хотим создать.

        workbook = xlsxwriter.Workbook('otchet_po_detyam123456.xlsx')
        worksheet = workbook.add_worksheet()
        data = [
            ['100', 'больше 90', 'больше 80', 'больше 70', 'не прошли порог'],
            [s, d, v, sm, n],
        ]
        for i in range(len(data[0])):
            worksheet.write('A' + str(i + 1), data[0][i])
            worksheet.write('B' + str(i + 1), data[1][i])
        chart1 = workbook.add_chart({'type': 'pie'})
        chart1.add_series({

            'name': 'Отчёт',

            'categories': ['Sheet1', 0, 0, len(data[0]) - 1, 0],

            'values': ['Sheet1', 0, 1, len(data[1]) - 1, 1],

        })
        chart1.set_title({'name': 'Отчёт'})
        chart1.set_style(10)
        worksheet.insert_chart('C2', chart1, {'x_offset': 25, 'y_offset': 10})
        workbook.close()
        self.make_image(all)

    def make_image(self, all):
        input_file = os.path.abspath('otchet_po_detyam123456.xlsx')
        output_image = 'otchet_po_detyam123456.png'
        operation = win32com.client.Dispatch("Excel.Application")
        operation.Visible = 0
        operation.DisplayAlerts = 0
        workbook_2 = operation.Workbooks.Open(input_file)
        sheet_2 = operation.Sheets(1)
        for x, chart in enumerate(sheet_2.Shapes):
            chart.Copy()
            image = ImageGrab.grabclipboard()
            image.save(output_image, 'png')
            pass
        workbook_2.Close(True)
        self.make_word(all)

    def make_word(self, all):
        workbook = xl.load_workbook(os.path.abspath('otchet_po_detyam123456.xlsx'))
        sheet_1 = workbook['Sheet1']
        template = DocxTemplate('template.docx')

        # Generate list of random values
        table_contents = []
        table_contents.append({
            'b100': sheet_1.cell(1, 2).value,
            'more90': sheet_1.cell(2, 2).value,
            'more80': sheet_1.cell(3, 2).value,
            'more70': sheet_1.cell(4, 2).value,
            'not': sheet_1.cell(5, 2).value
        })

        # Import saved figure
        image = InlineImage(template, 'otchet_po_detyam123456.png', Cm(10))

        # Declare template variables
        context = {
            'num': all,
            'day': datetime.datetime.now().day,
            'month': datetime.datetime.now().month,
            'year': datetime.datetime.now().year,
            'table_contents': table_contents,
            'image': image
        }

        # Render automated report
        template.render(context)
        template.save((self.lineEdit.text().split('.'))[0] + '.docx')
        os.remove(os.path.abspath('otchet_po_detyam123456.xlsx'))
        os.remove(os.path.abspath('otchet_po_detyam123456.png'))

    def run(self):
        global kod
        fname = QFileDialog.getOpenFileName(
            self, 'Выбрать таблицу для фильтровки', '',
            'Таблица (*.xlsx)')[0]
        workbook = load_workbook(fname)
        ws = workbook.active
        n = 0
        s = []
        k = 0
        n2 = 0
        f = 0
        kl = 0
        b = 0
        r = 0
        x = 0
        for row in ws.values:
            kod = 0
            klass = 0
            name = 0
            family = 0
            name2 = 0
            ball = 0
            r += 1
            for value in row:
                if value and (type(value) is str):
                    if 'Минимальный порог' in value:
                        porog = value[(value.find('Минимальный порог - ') + 20):]
                        porog = porog[:porog.find(' и')]
                        prg = ''
                        for j in porog:
                            if j != ' ':
                                prg += j
                        prg = float(prg)
                if k != 1:
                    kod += 1
                if n != 1:
                    name += 1
                if n2 != 1:
                    name2 += 1
                if f != 1:
                    family += 1
                if kl != 1:
                    klass += 1
                if b != 1:
                    ball += 1
                if value == 'Код ОО':
                    k = 1
                elif value == 'Класс':
                    kl = 1
                elif value == 'Фамилия':
                    f = 1
                elif value == 'Имя':
                    n = 1
                elif value == 'Отчество':
                    n2 = 1
                elif value == 'Балл':
                    b = 1
                if (n == 1) and (n2 == 1) and (k == 1) and (kl == 1) and (f == 1) and (b == 1):
                    break
            if (n == 1) and (n2 == 1) and (k == 1) and (kl == 1) and (f == 1) and (b == 1):
                break
        q = []
        for row in ws.iter_rows(min_row=r):
            t = []
            r1 = []
            for i in row:
                r1.append(i.value)
            if r1[kod - 1] == int(self.lineEdit_2.text()):
                t.append(r1[family - 1])
                t.append(r1[name - 1])
                t.append(r1[name2 - 1])
                t.append(r1[klass - 1])
                t.append(r1[ball - 1])
                if x != 1:
                    s.append(['Фамилия', [r1[family - 1]]])
                    s.append(['Имя', [r1[name - 1]]])
                    s.append(['Отчество', [r1[name2 - 1]]])
                    s.append(['Класс', [r1[klass - 1]]])
                    s.append(['Балл', [r1[ball - 1]]])
                else:
                    s[0][1].append(r1[family - 1])
                    s[1][1].append(r1[name - 1])
                    s[2][1].append(r1[name2 - 1])
                    s[3][1].append(r1[klass - 1])
                    s[4][1].append(r1[ball - 1])
                x = 1
            if (r1[kod - 1] != int(self.lineEdit_2.text())) and (x == 1):
                break
            if t:
                q.append(t)
        workbook.close()
        wb2 = xlsxwriter.Workbook(self.lineEdit.text())
        ws2 = wb2.add_worksheet()
        o = len(s[0][1])
        sto = list(filter(lambda x: x[4] == 100, q))
        s.append(['Количество учащихся принявших участие в ГИА', [o]])
        if sto:
            s.append(['ФИО стобальников', []])
            for i in sto:
                s[-1][1].append(i[0] + ' ' + i[1] + ' ' + i[2])
        dev = list(filter(lambda x: x[4] >= 90, q))
        devl = len(dev)
        if dev:
            s.append(['количество набравших больше 90', []])
            s.append(['ФИО набравших больше 90', []])
            for i in dev:
                s[-1][1].append(i[0] + ' ' + i[1] + ' ' + i[2])
            s[-2][1].append(devl)
        vos = list(filter(lambda x: x[4] >= 80, q))
        vosl = len(vos)
        if vos:
            s.append(['количество набравших больше 80', []])
            s.append(['ФИО набравших больше 80', []])
            for i in vos:
                s[-1][1].append(i[0] + ' ' + i[1] + ' ' + i[2])
            s[-2][1].append(vosl)
        sem = list(filter(lambda x: x[4] >= 70, q))
        seml = len(sem)
        if sem:
            s.append(['количество набравших больше 70', []])
            s.append(['ФИО набравших больше 70', []])
            for i in sem:
                if i[2]:
                    s[-1][1].append(i[0] + ' ' + i[1] + ' ' + i[2])
                else:
                    s[-1][1].append(i[0] + ' ' + i[1])
            s[-2][1].append(seml)
        sam = 0
        for i in q:
            sam += i[-1]
        sr = sam / len(q)
        s.append(['Средний балл по школе', [round(sr, 2)]])
        prc = list(filter(lambda x: x[-1] < prg, q))
        pr = len(prc)
        if prc:
            s.append(['Количество не преодолевших порог', []])
            s.append(['Не преодолели порог', []])
            for i in prc:
                if i[2]:
                    s[-1][1].append(i[0] + ' ' + i[1] + ' ' + i[2])
                else:
                    s[-1][1].append(i[0] + ' ' + i[1])
            s[-2][1].append(pr)
        for col, (name, value) in enumerate(s):
            ws2.write(0, col, name)
            p = 0
            for i in value:
                p += 1
                ws2.write(p, col, i)
        wb2.close()
        if self.lineEdit_3.text():
            z = [['Фамилия', []], ['Имя', []], ['Отчество', []], ['Балл', []]]
            q = list(filter(lambda x: x[3] == self.lineEdit_3.text(), q))
            for i in q:
                z[0][1].append(i[0])
                z[1][1].append(i[1])
                z[2][1].append(i[2])
                z[3][1].append(i[4])
            workbook = xlsxwriter.Workbook(self.lineEdit_4.text())
            ws = workbook.add_worksheet(self.lineEdit_3.text())
            sto = list(filter(lambda x: x[4] == 100, q))
            if sto:
                z.append(['ФИО стобальников', []])
                for i in sto:
                    z[-1][1].append(i[0] + ' ' + i[1] + ' ' + i[2])
            dev = list(filter(lambda x: x[4] >= 90, q))
            devl = len(dev)
            if dev:
                z.append(['количество набравших больше 90', []])
                z.append(['ФИО набравших больше 90', []])
                for i in dev:
                    z[-1][1].append(i[0] + ' ' + i[1] + ' ' + i[2])
                z[-2][1].append(devl)
            vos = list(filter(lambda x: x[4] >= 80, q))
            vosl = len(vos)
            if vos:
                z.append(['количество набравших больше 80', []])
                z.append(['ФИО набравших больше 80', []])
                for i in vos:
                    z[-1][1].append(i[0] + ' ' + i[1] + ' ' + i[2])
                z[-2][1].append(vosl)
            sem = list(filter(lambda x: x[4] >= 70, q))
            seml = len(sem)
            if sem:
                z.append(['количество набравших больше 70', []])
                z.append(['ФИО набравших больше 70', []])
                for i in sem:
                    if i[2]:
                        z[-1][1].append(i[0] + ' ' + i[1] + ' ' + i[2])
                    else:
                        z[-1][1].append(i[0] + ' ' + i[1])
                z[-2][1].append(seml)
            z.append(['Средний балл', [round((sum(z[3][1]) / len(z[3][1])), 2)]])
            pr = len(list(filter(lambda x: x[-1] >= prg, q)))
            z.append(['Преодолели порог', [pr]])
            for col, (name, value) in enumerate(z):
                ws.write(0, col, name)
                p = 0
                for i in value:
                    p += 1
                    ws.write(p, col, i)
            workbook.close()
        self.make_list(len(sto), devl, vosl, seml, pr, o)
        self.close()


def except_hook(cls, exception, traceback):
    sys.__excepthook__(cls, exception, traceback)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = Example()
    ex.show()
    sys.excepthook = except_hook
    sys.exit(app.exec())
